from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.tablefind.table_resolve import WorkbookConfig, SheetConfig, TablePlacement
from src.tablefind.table_models import StyleInfo
from typing import Dict, Any

def apply_style(cell: Any, style: StyleInfo):
    if style.font_name or style.font_size or style.bold or style.italic or style.underline:
        font = Font(
            name=style.font_name,
            size=style.font_size,
            bold=style.bold,
            italic=style.italic,
            underline='single' if style.underline else None
        )
        cell.font = font
    
    if style.font_color:
        cell.font = Font(color=style.font_color)
    
    if style.background_color:
        cell.fill = PatternFill(start_color=style.background_color, end_color=style.background_color, fill_type="solid")

def style_to_openpyxl(style: StyleInfo) -> Dict[str, Any]:
    openpyxl_style = {}
    if style.font_name or style.font_size or style.bold or style.italic or style.underline or style.font_color:
        openpyxl_style['font'] = Font(
            name=style.font_name,
            size=style.font_size,
            bold=style.bold,
            italic=style.italic,
            underline='single' if style.underline else None,
            color=style.font_color
        )
    if style.background_color:
        openpyxl_style['fill'] = PatternFill(start_color=style.background_color, end_color=style.background_color, fill_type="solid")
    return openpyxl_style

def place_table(sheet: Any, table_placement: TablePlacement):
    table = table_placement.table
    start_cell = table_placement.start_cell
    start_row, start_col = int(start_cell[1:]), ord(start_cell[0]) - ord('A') + 1

    # Write headers
    for col, header in enumerate(table.headers, start=start_col):
        cell = sheet.cell(row=start_row, column=col, value=header.text)
        if header.style:
            apply_style(cell, header.style)

    # Write data
    for row_idx, row in enumerate(table.rows, start=start_row + 1):
        for col_idx, cell_data in enumerate(row.cells, start=start_col):
            cell = sheet.cell(row=row_idx, column=col_idx, value=cell_data.text)
            if cell_data.style:
                apply_style(cell, cell_data.style)

    # Auto-adjust column widths
    for col in range(start_col, start_col + len(table.headers)):
        sheet.column_dimensions[get_column_letter(col)].auto_size = True

def create_excel_workbook(config: WorkbookConfig) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    for sheet_config in config.sheets:
        sheet = wb.create_sheet(sheet_config.name)
        
        if sheet_config.style:
            sheet_style = style_to_openpyxl(sheet_config.style)
            for row in sheet.iter_rows():
                for cell in row:
                    for style_attr, style_value in sheet_style.items():
                        setattr(cell, style_attr, style_value)

        for table_placement in sheet_config.tables:
            place_table(sheet, table_placement)

    return wb

def save_workbook(workbook: Workbook, filename: str):
    workbook.save(filename)
